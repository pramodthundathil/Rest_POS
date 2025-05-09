from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger 
from django.http import JsonResponse
from django.template.loader import render_to_string


from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from Home.decorators import admin_only, allowed_users
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .forms import RestaurantDetailsForm

from Finance.models import Income, Expence
# Create your views here.

@login_required(login_url='SignIn')
def Add_Category(request):
    if request.method == "POST":
        pic = request.FILES['pic']
        cname = request.POST['cname']
        foodcategory = FoodCategory.objects.create(image = pic, name= cname)
        foodcategory.save()
        messages.success(request,"Food Category Addedd...")
        return redirect("List_Category")
    
    return render(request,'add-category.html')

@login_required(login_url='SignIn')
def List_Category(request):
    food_category = FoodCategory.objects.all()
    p = Paginator(food_category, 20)
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number)  # returns the desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj = p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    context = {
        "food_category":page_obj
    }
    return render(request,'list-category.html',context)

@login_required(login_url='SignIn')
def DeleteCategory(request,pk):
    cat = FoodCategory.objects.get(id = pk)
    cat.delete()
    messages.success(request,'Food Category Deleted')
    return redirect('List_Category')


@login_required(login_url='SignIn')
def AddTax(request):
    if request.method == "POST":
        name = request.POST.get('name')
        tax_rate = request.POST.get('tax')
        tax = Tax.objects.create(tax_name = name,tax_percentage = tax_rate )
        tax.save()
        messages.success(request,'Tax Value Added Success')
        return redirect("ListTax")
    return render(request,"add-tax-slab.html")

@login_required(login_url='SignIn')
def ListTax(request):
    tax = Tax.objects.all()

    context = {
        "tax":tax
    }
    return render(request,"list-tax.html",context)


@login_required(login_url='SignIn')
def Add_Product(request):
    food_category = FoodCategory.objects.all()
    tax = Tax.objects.all()
    description = " "
    code  = " "
    if request.method == "POST":
        name = request.POST['name']
        category = FoodCategory.objects.get(id = int(request.POST['category']))
        code = request.POST['code']
        potion = request.POST['potion']
        diet = request.POST['diet']
        price = request.POST['price']
        stock = request.POST['stock']
        image = request.FILES['pic']
        description = request.POST['description']
        tax_name = request.POST["tax_name"]
        tax_value = request.POST["tax_value"]
        
        if Menu.objects.filter(code = code).exists():
            messages.info(request, 'Menu Item with same code is already exists...')
            return redirect("Add_Product")
        else:

            menu = Menu.objects.create(
                name = name, 
                category = category, 
                image = image, 
                code  = code,
                potion = potion, 
                diet =diet, 
                price = price, 
                stock = stock, 
                description = description,
                tax = tax_name,
                tax_value = Tax.objects.get(id = int(tax_value))

                )
            menu.save()
            messages.success(request,"Menu Item added Success...")
            return redirect("List_Product")
    
    context = {
        "food_category":food_category,
        "tax":tax

    }
    return render(request,'add-product.html',context)

@login_required(login_url='SignIn')
def List_Product(request):
    menu = Menu.objects.all()

    context = {
        "menu":menu,
    }
    return render(request,'list-product.html',context)

@login_required(login_url='SignIn')
def EditProduct(request, pk):
   
    menu = get_object_or_404(Menu, id=pk)
    categories = FoodCategory.objects.all()
    taxes = Tax.objects.all()
    
    if request.method == 'POST':
        menu.name = request.POST.get('mname')
        menu.price = float(request.POST.get('price'))
        menu.potion = request.POST.get('potion')
        menu.category_id = request.POST.get('category')
        menu.stock = request.POST.get('stock')
        menu.tax = request.POST.get('tax')
        menu.tax_value_id = request.POST.get('tax_value')
        menu.code = request.POST.get("code")
        # Handle file upload
        image = request.FILES.get('image')
        if image:
            menu.image = image
        
        # Calculate tax
        if menu.tax_value:
            tax_rate = menu.tax_value.tax_percentage / 100
            if menu.tax == "Exclusive":
                menu.tax_amount = round(menu.price * tax_rate, 2)
                menu.price_Before_tax = round(menu.price, 2)
                menu.price = round(menu.price + menu.tax_amount, 2)
            elif menu.tax == "Inclusive":
                menu.price_Before_tax = round(menu.price / (1 + tax_rate), 2)
                menu.tax_amount = round(menu.price - menu.price_Before_tax, 2)
        else:
            menu.price_Before_tax = round(menu.price, 2)
            menu.tax_amount = 0.0
        
        try:
            menu.save()
            messages.success(request, 'Menu updated successfully')
            return redirect('EditProduct',pk = pk)  # Replace with the name of the view you want to redirect to
        except Exception as e:
            messages.error(request, f'Error updating menu: {e}')
    
    context = {
        'menu': menu,
        'categories': categories,
        'taxes': taxes,
    }
    return render(request, "edit-product.html", context)


@login_required(login_url='SignIn')
def EditCategory(request,pk):

    cat = get_object_or_404(FoodCategory, id=pk)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        image = request.FILES.get('image')
        
        if name:
            cat.name = name
        
        if image:
            cat.image = image
        
        try:
            cat.save()
            messages.success(request, 'Category updated successfully')
            return redirect('EditCategory', pk = pk)  
        except Exception as e:
            messages.error(request, f'Error updating category: {e}')
    
    context = {
        'cat': cat,
    }
    return render(request,'edit-category.html',context)

@login_required(login_url='SignIn')
def DeleteProduct(request,pk):
    menu  = Menu.objects.get(id = pk)
    if menu.status == False:
        menu.status = True
    else:
        menu.status = False
    menu.save()
    messages.info(request,"Product Deleted....")
    return redirect("List_Product")


@login_required(login_url='SignIn')
def Add_Table(request):
    if request.method == "POST":
        tnum = request.POST['tnum']
        seats = request.POST['seats']
        if Tables.objects.filter(Table_number = tnum).exists():
            messages.error(request,"Table Already Exists...")
            return redirect("Add_Table")
        else:
            table = Tables.objects.create(Table_number = tnum, Number_of_Seats = seats)
            table.save()
            messages.success(request,"Table added Success...")
            return redirect("List_Table")

    return render(request,"add-table.html")

@login_required(login_url='SignIn')
def List_Table(request):
    table = Tables.objects.all()
    context = {
        "table":table
    }
    return render(request,"list-table.html",context)

@login_required(login_url='SignIn')
def Delete_Table(request,pk):
    Tables.objects.get(id = pk).delete()
    messages.success(request,"Table deleted success....")
    return redirect("List_Table")

@login_required(login_url='SignIn')
def Pos(request):
    category = FoodCategory.objects.all()
    menu = Menu.objects.filter(status = True)
    table = Tables.objects.all()
    orders = Order.objects.filter(user = request.user, checkout_status = False)
    order_details = []

    for order in orders:
        total_items = sum(item.quantity for item in order.items.all())
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_details.append({
            'order': order,
            'total_items': total_items,
            'total_price': total_price,
        })

    in_progress_orders = Order.objects.filter(completion_status=False,checkout_status = False, status="In Progress")
    preparing_orders = Order.objects.filter(completion_status=False,checkout_status = False, status="In Kitchen")
    pending_orders = Order.objects.filter(completion_status=False,checkout_status = False, status="Pending")

    # Combine results if needed
    all_orders = Order.objects.filter(completion_status=False,checkout_status = False,status__in=["In Progress", "In Kitchen", "Pending"])

    # Display count
    in_progress_count = in_progress_orders.count()
    preparing_count = preparing_orders.count()
    pending_count = pending_orders.count()
    # Combine results if needed
    all_orders = Order.objects.filter(completion_status=False,checkout_status = False, status__in=["In Progress", "In Kitchen", "Pending"])
    print(all_orders,'------------------------------------------')
    context = {
        "category":category,
        "menu":menu,
        "table":table,
        "orders":orders,
        'order_details': order_details,
        "in_progress_orders": in_progress_orders,
        "preparing_orders": preparing_orders,
        "pending_orders": pending_orders,
        "in_progress_count":in_progress_count,
        "preparing_count":preparing_count,
        "pending_count":pending_count,
        "all_orders": all_orders.count()
    }
    return render(request,'posinterface.html',context)

@login_required(login_url='SignIn')
@admin_only
def PosIndex(request):
    category = FoodCategory.objects.filter()
    menu = Menu.objects.filter(status = True)
    table = Tables.objects.all()
    orders = Order.objects.filter(checkout_status = False)
    order_details = []

    in_progress_orders = Order.objects.filter(completion_status=False,checkout_status = False, status="In Progress")
    preparing_orders = Order.objects.filter(completion_status=False,checkout_status = False, status="In Kitchen")
    pending_orders = Order.objects.filter(completion_status=False,checkout_status = False, status="Pending")

    # Combine results if needed
    all_orders = Order.objects.filter(completion_status=False,checkout_status = False, status__in=["In Progress", "In Kitchen", "Pending"])

    # Display count
    in_progress_count = in_progress_orders.count()
    preparing_count = preparing_orders.count()
    pending_count = pending_orders.count()
    all_count = all_orders.count()

    for order in orders:
        total_items = sum(item.quantity for item in order.items.all())
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_details.append({
            'order': order,
            'total_items': total_items,
            'total_price': total_price,
            "in_progress_count":in_progress_count,
            "preparing_count":preparing_count,
            "pending_count":pending_count,
            "all_orders":all_orders.count()

        })

    context = {
        "category":category,
        "menu":menu,
        "table":table,
        "orders":orders,
        'order_details': order_details,
        
    }
    return render(request,'posinterface.html',context)

@login_required(login_url='SignIn')
def OrderSingle(request,pk):
    category = FoodCategory.objects.all()
    menu = Menu.objects.filter(status = True)
    order = Order.objects.get(id = pk)
    item = OrderItem.objects.filter(order = order)
    total_price = round(sum(item.get_total_price() for item in order.items.all()),2)
    addons = AddOns.objects.all()
    users = User.objects.filter(groups__name = "dboy")

    


    context = {
        "category":category,
        "menu":menu,
        "order":order,
        "item":item,
        "total_price":round(total_price,2),
        "addons":addons,
        "users":users
    }
    return render(request,"order-single.html",context)


@login_required(login_url='SignIn')
def search_menu(request):
    query = request.GET.get('product_search', '')
    menu = Menu.objects.filter(status = True)
    order_id = request.GET.get('order_id', '')

    category = FoodCategory.objects.all()
    order = Order.objects.get(id = int(order_id))
    addons = AddOns.objects.all()



    if query:
        results = Menu.objects.filter(name__icontains=query) | Menu.objects.filter(code__icontains=query)
    else:
        results = Menu.objects.none()
    
    return render(request, 'search_results.html', {'results': results,"category":category,"menu":menu,"order":order,"addons":addons})

@login_required(login_url='SignIn')
def CreateOrder(request):
    if request.method == "POST":
        table = Tables.objects.get(id = int(request.POST['table']))
        order = Order.objects.create(table = table,user = request.user )
        order.save()
        
        return redirect('OrderSingle',pk = order.id )
    


@login_required(login_url='SignIn')
def add_to_order(request):
    if request.method == "POST":
        menu_item_id = request.POST.get('menu_item_id')
        order_id = request.POST.get('order_id')
        menu_item = get_object_or_404(Menu, id=menu_item_id)
        order = get_object_or_404(Order, id=order_id)
        addons = AddOns.objects.all()
        users = User.objects.filter(groups__name = "dboy")


        # Create or update the OrderItem
        order_item, created = OrderItem.objects.get_or_create(order=order, menu_item=menu_item, defaults={'price': menu_item.price})
        if not created:
            order_item.quantity += 1
            order_item.save()

        # Render the updated order items and return as HTML
        item = OrderItem.objects.filter(order = order)
        order.total_price = sum(item.get_total_price() for item in order.items.all())
        order.save()
        total_price = order.total_price
        order_html = render_to_string('order-summery.html', {'order': order,"item":item,"total_price":total_price,"addons":addons,"users":users})

        return JsonResponse({'order_html': order_html})
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required(login_url='SignIn')
@csrf_exempt
def add_on_to_item(request,pk,id):
    if request.method == "POST":
        order_item = OrderItem.objects.get(id = pk)  # Assuming a function to get the current order

        special_instructions = request.POST.get('instraction', '')
        addon_ids = request.POST.getlist('addons')
        print(addon_ids,"---------------------")
        try:
            addon_ids = [int(addon_id) for addon_id in addon_ids]
        except ValueError:
            print("Invalid add-on IDs provided.")
            messages.error(request, "Invalid add-ons selected.")
            return redirect('OrderSingle', pk=id)

        addons = AddOns.objects.filter(id__in=addon_ids)
        print("Addons queryset:", addons)

        order_item.add_ons.set(addons)
        order_item.special_instructions = special_instructions
        order_item.save()
        messages.info(request,"Add on added ")

    return redirect('OrderSingle',pk = id)

@login_required(login_url='SignIn')
def add_items_to_order(request,pk,id):
    return redirect('OrderSingle',pk = id)

@login_required(login_url='SignIn')
def increase_quantity(request):
    if request.method == "POST":
        item_id = request.POST.get('item_id')
        order_item = get_object_or_404(OrderItem, id=item_id)
        order_item.quantity += 1
        order_item.save()
        users = User.objects.filter(groups__name = "dboy")

        item = OrderItem.objects.filter(order = order_item.order)
        total_price = sum(item.get_total_price() for item in order_item.order.items.all())
        order_html = render_to_string('order-summery.html', {'order': order_item.order, 'total_price': total_price,"item":item,"users":users})

        return JsonResponse({'order_html': order_html})
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required(login_url='SignIn')
def decrease_quantity(request):
    users = User.objects.filter(groups__name = "dboy")
    if request.method == "POST":
        item_id = request.POST.get('item_id')
        order_item = get_object_or_404(OrderItem, id=item_id)
        if order_item.quantity > 1:
            order_item.quantity -= 1
            order_item.save()

        item = OrderItem.objects.filter(order = order_item.order)
        total_price = sum(item.get_total_price() for item in order_item.order.items.all())
        order_html = render_to_string('order-summery.html', {'order': order_item.order, 'total_price': total_price,"item":item,"users":users})

        return JsonResponse({'order_html': order_html})
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required(login_url='SignIn')
def Delete_menuitem(request,pk):
    item = OrderItem.objects.get(id = pk)
    order = item.order.id
    item.delete()
    return redirect("OrderSingle",pk = order)

@login_required(login_url='SignIn')
@csrf_exempt
def TakeOrder(request, pk):
    order = get_object_or_404(Order, id=pk)
    if request.method == "POST":
        vehicle_number = request.POST.get('vehicle_number', "")
        delivery_boy_id = request.POST.get("dboy")
        
        # Update delivery boy if provided
        if delivery_boy_id:
            try:
                dboy = get_object_or_404(User, id=int(delivery_boy_id))
                order.delivery_boy = dboy
                print(f"Delivery boy set to: {dboy.first_name}")  # Debug log
            except Exception as e:
                print(f"Error setting delivery boy: {e}")  # Debug log
        
        # Update vehicle number if provided
        if vehicle_number:
            order.vehicle_number = vehicle_number
            print(f"Vehicle number set to: {vehicle_number}")  # Debug log
            
        # Update order status
        order.take_order = True
        order.completion_status = False
        order.save()
        
        # Send update through WebSocket
        try:
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "updates",
                {
                    "type": "send_update",
                    "message": "Database updated",
                }
            )
        except Exception as e:
            print(f"Error sending WebSocket update: {e}")
        
        # For AJAX requests, return JSON response
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Order placed successfully'})
        
        # For regular requests, redirect
        return redirect("Pos")
    else:
        # Handle GET request
        return redirect("Pos")


    
# from escpos.printer import Usb
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

import os
# import tempfile

# from weasyprint import HTML

# @login_required(login_url='SignIn')
# def print_invoice(request, order_id):
#     order = get_object_or_404(Order, id=order_id)
#     rest_details = RestaurantDetails.objects.all().last()
#     items = order.items.all()
#     total_price = sum(item.get_total_price() for item in items)
#     context = {
#         'order': order,
#         'item': items,
#         'total_price': round(total_price,2),
#         "rest_details":rest_details
#     }

#     # Render the HTML template
#     html_content = render_to_string('receipt.html', context)

#     # Convert HTML to PDF using WeasyPrint
#     pdf_file = HTML(string=html_content).write_pdf()

#     pdf_file = HTML(string=html_content).write_pdf()

#     # Use tempfile to create a temporary file
#     with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
#         temp_pdf.write(pdf_file)
#         temp_pdf_path = temp_pdf.name

#     # Send the PDF to the printer
#     printer_name = "EPSON L3210 Series"
#     printer_handle = win32print.OpenPrinter(printer_name)

#     # Open the file to print
#     with open(temp_pdf_path, 'rb') as f:
#         pdf_data = f.read()

#     # Use win32print to print the file
#     win32print.StartDocPrinter(printer_handle, 1, ("Invoice", None, "RAW"))
#     win32print.StartPagePrinter(printer_handle)
#     win32print.WritePrinter(printer_handle, pdf_data)
#     win32print.EndPagePrinter(printer_handle)
#     win32print.EndDocPrinter(printer_handle)
#     win32print.ClosePrinter(printer_handle)

#     return HttpResponse("Invoice printed successfully")




import os
import tempfile
# from weasyprint import HTML
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

import os
# import subprocess
import tempfile
# import win32print
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
# from weasyprint import HTML

@login_required(login_url='SignIn')
def print_invoice(request, order_id):
    # order = get_object_or_404(Order, id=order_id)
    # rest_details = RestaurantDetails.objects.all().last()
    # items = order.items.all()

    # total_price = sum(item.get_total_price() for item in items)
    # vat_rate = 0.05  # 5% VAT
    # vat_amount = round(total_price * vat_rate, 2)
    # grand_total = round(total_price + vat_amount, 2)

    # # Create the invoice in HTML format
    # receipt_html = f"""
    # <h2>{rest_details.Name_of_restaurant}</h2>
    # <p>{rest_details.location}</p>
    # <p><strong>Tel:</strong> {rest_details.phone}</p>
    # <p><strong>TRN:</strong> {rest_details.TRN}</p>

    # <h3>TAX INVOICE</h3>
    # <p><strong>Bill#:</strong> {order.id}</p>
    # <p><strong>Date:</strong> {order.create_date}</p>
    # <p><strong>Table No:</strong> Takeaway</p>

    # <hr>
    # <table width='100%' border='1' cellspacing='0' cellpadding='5'>
    #     <tr>
    #         <th>Description</th>
    #         <th>Qty</th>
    #         <th>Amount</th>
    #     </tr>
    # """

    # for item in items:
    #     receipt_html += f"""
    #     <tr>
    #         <td>{item.menu_item.name}</td>
    #         <td>{item.quantity}</td>
    #         <td>DHS {item.get_total_price():.2f}</td>
    #     </tr>
    #     """

    # receipt_html += f"""
    # </table>
    # <hr>
    # <p><strong>Total:</strong> DHS {total_price:.2f}</p>
    # <p><strong>VAT (5%):</strong> DHS {vat_amount:.2f}</p>
    # <p><strong>Grand Total:</strong> DHS {grand_total:.2f}</p>

    # <p><strong>Payment Method:</strong> {order.payment_status}</p>
    # <p><strong>Amount Paid:</strong> DHS {order.total_price}</p>

    # <p><strong>Cashier Name:</strong> admin</p>
    # <p>*** THANK YOU, COME AGAIN ***</p>
    # """

    # # Generate PDF
    # pdf_file = HTML(string=receipt_html).write_pdf()

    # # Save to a temporary file
    # with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
    #     temp_pdf.write(pdf_file)
    #     temp_pdf_path = temp_pdf.name

    # try:
    #     # Use Adobe Acrobat Reader (Modify this path if needed)
    #     acrobat_path = r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe"
        
    #     if not os.path.exists(acrobat_path):
    #         return HttpResponse("Adobe Acrobat Reader not found", status=500)

    #     # Print using Adobe Acrobat silently
    #     subprocess.run([acrobat_path, "/p", "/h", temp_pdf_path], check=True)
        
    #     return HttpResponse("Invoice sent to printer successfully")

    # except Exception as e:
    #     return HttpResponse(f"Printer error: {str(e)}", status=500)

    # finally:
    #     # Clean up temp file
    #     if os.path.exists(temp_pdf_path):
    #         os.remove(temp_pdf_path)
    return redirect("Pos")


# @login_required(login_url='SignIn')
# def print_invoice(request, order_id):
#     order = get_object_or_404(Order, id=order_id)
#     rest_details = RestaurantDetails.objects.all().last()
#     items = order.items.all()
#     total_price = sum(item.get_total_price() for item in items)
#     context = {
#         'order': order,
#         'item': items,
#         'total_price': round(total_price, 2),
#         "rest_details": rest_details
#     }

#     # Render the HTML template
#     html_content = render_to_string('receipt.html', context)

#     # Convert HTML to PDF using WeasyPrint
#     pdf_file = HTML(string=html_content).write_pdf()

#     # Use tempfile to create a temporary file
#     with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
#         temp_pdf.write(pdf_file)
#         temp_pdf_path = temp_pdf.name

#     # Use Adobe Acrobat Reader or SumatraPDF to send the PDF to the printer
#     printer_name = "EPSON L3210 Series"  # Replace with your printer's name
    
#     # Set the PDF application you want to use to print
#     # For Adobe Acrobat Reader (if installed):
#     os.system(f'"C:\\Program Files\\Adobe\\Acrobat Reader DC\\Acrobat\\Acrobat.exe" /t "{temp_pdf_path}" "{printer_name}"')
    
#     # For SumatraPDF (if installed):
#     # os.system(f'"C:\\Program Files\\SumatraPDF\\SumatraPDF.exe" -print-to "{printer_name}" "{temp_pdf_path}"')

#     return HttpResponse("Invoice printed successfully")




@login_required(login_url='SignIn')
def receipt_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    rest_details = RestaurantDetails.objects.all().last()
    items = order.items.all()
    total_price = sum(item.get_total_price() for item in items)
    context = {
        'order': order,
        'item': items,
        'total_price': round(total_price,2),
        "rest_details":rest_details
    }
    return render(request, 'receipt.html', context)


@login_required(login_url='SignIn')
def KitchenDashboard(request):
    orders = Order.objects.filter( take_order= True, completion_status = False )
    order_details = []
    orderitem = OrderItem.objects.all()
    for order in orders:
        orderitem = OrderItem.objects.filter(order = order)
        order_details.append(
            {
                "order":order,
                "orderitem":orderitem
            }
        ) 

    context = {
        "order_details":order_details
    } 

    return render(request,"kitchendash.html",context)



def refresh_table(request):
    orders = Order.objects.filter(take_order = True,completion_status = False, checkout_status = False)
    order_details = []
    orderitem = OrderItem.objects.all()
    for order in orders:
        orderitem = OrderItem.objects.filter(order = order)
        order_details.append(
            {
                "order":order,
                "orderitem":orderitem
            }
        ) 

    context = {
        "order_details":order_details
    }
    table_html = render_to_string('kitchendashitems.html', context)
    return JsonResponse({'table_html': table_html})

def refresh_order(request):
    category = FoodCategory.objects.filter()
    menu = Menu.objects.filter(status = True)
    table = Tables.objects.all()
    orders = Order.objects.filter(checkout_status = False)
    order_details = []

    for order in orders:
        total_items = sum(item.quantity for item in order.items.all())
        total_price = sum(item.get_total_price() for item in order.items.all())
        order_details.append({
            'order': order,
            'total_items': total_items,
            'total_price': total_price,

        })

    context = {
        "category":category,
        "menu":menu,
        "table":table,
        "orders":orders,
        'order_details': order_details,
        
    }
    table_html = render_to_string('order-datas.html', context)
    return JsonResponse({'table_html': table_html})

def Status_Change(request):
    if request.method == "POST":
        order_id = request.POST.get('orderid')  # Corrected key name
        order = get_object_or_404(Order, id=order_id)
        order.status = "In Progress"
        order.save()
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
        )
        
        orders = Order.objects.filter(take_order=True,completion_status = False, checkout_status = False)
        order_details = []
        for order in orders:
            orderitem = OrderItem.objects.filter(order=order)
            order_details.append({
                "order": order,
                "orderitem": orderitem
            }) 

        context = {
            "order_details": order_details
        }
        table_html = render_to_string('kitchendashitems.html', context)
        return JsonResponse({'order_html': table_html})
    
def Status_Change_Order_Ready(request):
    if request.method == "POST":
        order_id = request.POST.get('orderid')  # Corrected key name
        order = get_object_or_404(Order, id=order_id)
        order.status = "Order Ready"
        order.save()
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
        )
        
        orders = Order.objects.filter(take_order=True,completion_status = False, checkout_status = False)
        order_details = []
        for order in orders:
            orderitem = OrderItem.objects.filter(order=order)
            order_details.append({
                "order": order,
                "orderitem": orderitem
            }) 

        context = {
            "order_details": order_details
        }
        table_html = render_to_string('kitchendashitems.html', context)
        return JsonResponse({'order_html': table_html})
    

def Status_Change_Menu_Finish(request):
    if request.method == "POST":
        order_item_id = request.POST.get('orderid')  # Ensure the correct key is used
        order_item = get_object_or_404(OrderItem, id=order_item_id)
        order_item.completion_status = True
        order_item.save()
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
        "updates",
        {
            "type": "send_update",
            "message": "Database updated",
            
        }
        )
        
        orders = Order.objects.filter(take_order=True,completion_status = False, checkout_status = False)
        order_details = []
        for order in orders:
            orderitem = OrderItem.objects.filter(order=order)
            order_details.append({
                "order": order,
                "orderitem": orderitem
            })

        context = {
            "order_details": order_details
        }
        table_html = render_to_string('kitchendashitems.html', context)
        return JsonResponse({'order_html': table_html})

    

def Status_Change_OrderCompeletion(request,pk):
    
    order = get_object_or_404(Order, id=pk)
    order.completion_status = True
    order.take_order = False
    order.save()
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
    "updates",
    {
        "type": "send_update",
        "message": "Database updated",
        
    }
    )
    return redirect("KitchenDashboard")

def calculate_tax(menu_item, quantity):
    tax_value = 0
    tax_value = menu_item.tax_amount * quantity
    return round(tax_value,2)

def SettleOrder(request, pk):
    try:
        order = get_object_or_404(Order, id=pk)
    except:
        messages.info(request,"Order not Found.....")
        return redirect("Pos")
    if request.method == "POST":
        payment = request.POST.get("payment")
        order_items = OrderItem.objects.filter(order=order)
        total_price = sum(item.menu_item.price * item.quantity for item in order_items)
        total_tax_amount = sum(calculate_tax(item.menu_item, item.quantity) for item in order_items)
        try:
            checkout = Checkout.objects.create(
                order=order,
                payment_method=payment,
                payment_status="Paid",
                total_price=total_price,
                tax_amount=total_tax_amount
            )
            checkout.save()
            order.checkout_status = True
            order.payment_method = payment
            order.payment_status = 'Paid'
            order.save()
            income = Income.objects.create(perticulers = f"payment Order number #{order.id} by {payment}", amount = total_price )
            income.save()
            messages.info(request, "Bill Settled....")

        except:
            messages.info(request, "This Bill Already settled")
        return redirect("Pos")

    # Render the order settlement page if it's a GET request
    return render(request, 'settle_order.html', {'order': order})
    
@allowed_users(allowed_roles=["admin"])   
def ViewCheckouts(request):
    checkout = Checkout.objects.all().order_by('-id')
    context = {
        "checkout":checkout
    }
    return render(request,"settledorders.html",context)

def delete_settled_order(request,pk):
    checkout = get_object_or_404(Checkout, id= pk)
    checkout.delete()
    messages.success(request, "order deleted....")
    return redirect("ViewCheckouts")


def Reports(request):
    return render(request,"reports.html")



from django.http import HttpResponse
from django.utils.timezone import now
import openpyxl
from openpyxl.styles import Font
from .models import Order, Checkout

def generate_excel_report(request):
    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Today\'s Orders Report'

    # Define the columns
    columns = ['Order ID', 'Order Date', 'Table', 'Total Price', 'Tax Amount', 'Payment Method', 'Payment Status']

    # Set the header row
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Fetch today's orders
    today = now().date()
    orders = Order.objects.filter(create_date__date=today)

    # Add the order data to the worksheet
    for order in orders:
        checkout = Checkout.objects.filter(order=order).first()
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = order.id
        worksheet.cell(row=row_num, column=2).value = order.create_date.strftime('%Y-%m-%d %H:%M')
        worksheet.cell(row=row_num, column=3).value = order.table.Table_number if order.table else 'N/A'
        worksheet.cell(row=row_num, column=4).value = checkout.total_price if checkout else 'N/A'
        worksheet.cell(row=row_num, column=5).value = checkout.tax_amount if checkout else 'N/A'
        worksheet.cell(row=row_num, column=6).value = checkout.payment_method if checkout else 'N/A'
        worksheet.cell(row=row_num, column=7).value = checkout.payment_status if checkout else 'N/A'

    # Set the column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=todays_orders_report.xlsx'
    workbook.save(response)
    return response

def generate_orders_report(request):
    if request.method == "POST":
        # Get start date and end date from the request
        start_date = request.POST.get('sdate')
        end_date = request.POST.get('edate')

        # Create a new Excel workbook and add a worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Orders Report'

        # Define the columns
        columns = ['Order ID', 'Order Date', 'Table', 'Total Price', 'Tax Amount', 'Payment Method', 'Payment Status']

        # Set the header row
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        # Fetch orders within the date range
        orders = Order.objects.filter(create_date__gte=start_date, create_date__lte=end_date)

        # Add the order data to the worksheet
        for order in orders:
            checkout = Checkout.objects.filter(order=order).first()
            row_num += 1
            worksheet.cell(row=row_num, column=1).value = order.id
            worksheet.cell(row=row_num, column=2).value = order.create_date.strftime('%Y-%m-%d %H:%M')
            worksheet.cell(row=row_num, column=3).value = order.table.Table_number if order.table else 'N/A'
            worksheet.cell(row=row_num, column=4).value = checkout.total_price if checkout else 'N/A'
            worksheet.cell(row=row_num, column=5).value = checkout.tax_amount if checkout else 'N/A'
            worksheet.cell(row=row_num, column=6).value = checkout.payment_method if checkout else 'N/A'
            worksheet.cell(row=row_num, column=7).value = checkout.payment_status if checkout else 'N/A'

        # Set the column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=orders_report_{start_date}_to_{end_date}.xlsx'
        workbook.save(response)
        return response
    

# extra add on for food menu 


def list_add_ons(request):
    menu = AddOns.objects.all()
    context = {
        "menu":menu
    }
    return render(request,"list-addons.html",context)


def add_add_ons(request):
    if request.method == "POST":
        name = request.POST.get('name')
        price = request.POST.get('price')
        addons = AddOns.objects.create(name = name, price = price)
        addons.save()
        messages.info(request,"Item Added.........")
        return redirect("list_add_ons")

    return render(request,"add-addons.html")
    






        
        

    



from .forms import RestaurantDetailsForm
from .models import RestaurantDetails
# resurgent Details adding

def profile(request):
    details = RestaurantDetails.objects.all().last()
    
    if request.method == "POST":
        if details:
            form = RestaurantDetailsForm(request.POST, request.FILES, instance=details)
        else:
            form = RestaurantDetailsForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            return redirect('profile')  # Redirect to the same page or any other page
    else:
        form = RestaurantDetailsForm(instance=details)

    context = {
        "form": form
    }
    return render(request, "user_profile.html", context)



# newly added edit functionalities 
from .forms import TablesForm, TaxForm, MenuForm, FoodCategoryForm, AddOnsForm

def edit_tax(request, pk):
    tax = get_object_or_404(Tax, id = pk)
    form = TaxForm(instance=tax)
    if request.method == "POST":
        form = TaxForm(request.POST, instance=tax)
        if form.is_valid():
            form.save()
            messages.success(request,"Tax updated..")
            return redirect(ListTax)
        else:
            messages.error(request,"Tax not updated..")
            return redirect(ListTax)
    return render(request,"edit_tax.html",{"form":form})

def delete_tax(request, pk):
    tax = get_object_or_404(Tax, id = pk)
    tax.delete()
    messages.success(request,"Tax Deleted..")
    return redirect(ListTax)

def edit_table(request, pk):
    table = get_object_or_404(Tables, id=pk)
    form = TablesForm(instance=table)
    if request.method == "POST":
        form = TablesForm(request.POST, instance=table)
        if form.is_valid():
            form.save()
            messages.success(request, "Table updated successfully.")
            return redirect('List_Table')
        else:
            messages.error(request, "Failed to update table.")
    context = {
        'form': form,
        'table': table,
    }
    return render(request, 'edit_table.html', context)

def delete_table(request, pk):
    table = get_object_or_404(Tables, id=pk)
    table.delete()
    messages.success(request, "Table deleted successfully.")
    return redirect('List_Table')

def edit_add_on(request, pk):
    add_on = get_object_or_404(AddOns, id=pk)
    form = AddOnsForm(instance=add_on)
    if request.method == "POST":
        form = AddOnsForm(request.POST, instance=add_on)
        if form.is_valid():
            form.save()
            messages.success(request, "Add-on updated successfully.")
            return redirect('list_add_ons')
        else:
            messages.error(request, "Failed to update add-on.")
    context = {
        'form': form,
        'add_on': add_on,
    }
    return render(request, 'edit_add_on.html', context)

def delete_add_on(request, pk):
    add_on = get_object_or_404(AddOns, id=pk)
    add_on.delete()
    messages.success(request, "Add-on deleted successfully.")
    return redirect('list_add_ons')

